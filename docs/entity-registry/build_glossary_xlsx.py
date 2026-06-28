#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

base = os.path.dirname(os.path.abspath(__file__))
ENT = [
 ("entities_core.csv","Ядро"),
 ("ggd_entities.csv","Гаура-ганоддеша-дипика · Гауранга Лила"),
 ("ggd_krishna_lila_entities.csv","Гаура-ганоддеша-дипика · Кришна Лила"),
 ("iskcon_gurus_entities.csv","Гуру ИСККОН (GBC)"),
 ("prabhupada_lilamrita_entities.csv","Прабхупада-лиламрита"),
 ("mahabharata_entities.csv","Махабхарата"),
 ("ramayana_entities.csv","Рамаяна"),
 ("dhama_entities.csv","Дхама"),
 ("extra_entities.csv","Гаудия-вайшнавы"),
 ("granthas_entities.csv","Грантхи ачарьев"),
]
REL = [
 ("relations_core.csv","Ядро"),
 ("ggd_relations.csv","Гаура-ганоддеша-дипика"),
 ("iskcon_gurus_relations.csv","Гуру ИСККОН (GBC)"),
 ("prabhupada_lilamrita_relations.csv","Прабхупада-лиламрита"),
 ("mahabharata_relations.csv","Махабхарата"),
 ("ramayana_relations.csv","Рамаяна"),
 ("book_relations.csv","Книги (appears-in/author)"),
 ("granthas_relations.csv","Грантхи ачарьев"),
 ("dhama_relations.csv","Дхама"),
]
REL_RU = {
 "avatar-of":"аватар (из)","expansion-of":"экспансия (из)","shaktyavesha-of":"шактьявеша (из)",
 "father-of":"отец","mother-of":"мать","son-of":"сын","foster-son-of":"приёмный сын",
 "brother-of":"брат","sister-of":"сестра","wife-of":"жена","husband-of":"муж",
 "grandson-of":"внук","nephew-of":"племянник","disciple-of":"ученик (кого)","guru-of":"гуру (кого)",
 "associate-of":"спутник","godbrother-of":"духовный брат","author-of":"автор",
 "speaker-of":"произнёс","hearer-of":"слушатель","narrator-of":"рассказчик",
 "appears-in":"встречается в","gauranga-lila-identity":"тождество в Гауранга Лиле",
}

# ---- load entities
ents=[]; by_id={}
for fn,label in ENT:
    for r in csv.DictReader(open(os.path.join(base,fn),encoding="utf-8")):
        r["dataset"]=label; ents.append(r); by_id[r["id"]]=r
def disp(r): return (r.get("name_ru") or r.get("name_en") or r["id"]).strip()
ents.sort(key=lambda r: disp(r).lower())

# ---- load relations
rels=[]
for fn,label in REL:
    for r in csv.DictReader(open(os.path.join(base,fn),encoding="utf-8")):
        r["dataset"]=label; rels.append(r)

# ---- master CSVs (canonical merge for repo / DB)
cols=["id","type","tattva","category","name_en","name_iast","name_ru","aliases","note","source_ref","confidence","iast_status","dataset"]
with open(os.path.join(base,"entities_all.csv"),"w",newline="",encoding="utf-8") as f:
    w=csv.DictWriter(f,fieldnames=cols); w.writeheader()
    for r in ents: w.writerow({k:r.get(k,"") for k in cols})
with open(os.path.join(base,"relations_all.csv"),"w",newline="",encoding="utf-8") as f:
    w=csv.writer(f); w.writerow(["from_id","relation","to_id","dataset"])
    for r in rels: w.writerow([r["from_id"],r["relation"],r["to_id"],r["dataset"]])

# ---- styling helpers
HEAD_FILL=PatternFill("solid",start_color="1F3864")
HEAD_FONT=Font(name="Arial",bold=True,color="FFFFFF",size=11)
BASE_FONT=Font(name="Arial",size=10)
THIN=Side(style="thin",color="D9D9D9")
BORDER=Border(bottom=THIN)
WRAP=Alignment(vertical="top",wrap_text=True)
TOP=Alignment(vertical="top")

def style_header(ws,ncol):
    for c in range(1,ncol+1):
        cell=ws.cell(row=1,column=c); cell.fill=HEAD_FILL; cell.font=HEAD_FONT
        cell.alignment=Alignment(vertical="center",horizontal="left")
    ws.freeze_panes="A2"
    ws.auto_filter.ref=f"A1:{get_column_letter(ncol)}{ws.max_row}"
    ws.row_dimensions[1].height=22

wb=Workbook()

# ===== Sheet 1: Glossary =====
ws=wb.active; ws.title="Глоссарий"
hdr=["№","id","Имя (RU)","Name (EN)","IAST","IAST статус","Таттва","Категория","Алиасы","Определение","Источник","Достоверность","Датасет"]
ws.append(hdr)
for i,r in enumerate(ents,1):
    ws.append([i,r["id"],r.get("name_ru",""),r.get("name_en",""),r.get("name_iast",""),r.get("iast_status",""),
               r.get("tattva",""),r.get("category",""),r.get("aliases",""),r.get("note",""),
               r.get("source_ref",""),r.get("confidence",""),r["dataset"]])
for row in ws.iter_rows(min_row=2,max_row=ws.max_row):
    for c in row: c.font=BASE_FONT; c.alignment=WRAP; c.border=BORDER
widths=[5,30,26,30,18,12,14,30,26,46,22,13,34]
for i,wd in enumerate(widths,1): ws.column_dimensions[get_column_letter(i)].width=wd
style_header(ws,len(hdr))

# ===== Sheet 2: Relations =====
ws2=wb.create_sheet("Связи")
hdr2=["№","Откуда (id)","Откуда (имя)","Связь","Связь (RU)","Куда (id)","Куда (имя)","Датасет"]
ws2.append(hdr2)
def name_of(eid): 
    r=by_id.get(eid); return (r and (r.get("name_ru") or r.get("name_en"))) or eid
for i,r in enumerate(rels,1):
    ws2.append([i,r["from_id"],name_of(r["from_id"]),r["relation"],REL_RU.get(r["relation"],""),
                r["to_id"],name_of(r["to_id"]),r["dataset"]])
for row in ws2.iter_rows(min_row=2,max_row=ws2.max_row):
    for c in row: c.font=BASE_FONT; c.alignment=TOP; c.border=BORDER
for i,wd in enumerate([5,26,26,22,22,26,26,30],1): ws2.column_dimensions[get_column_letter(i)].width=wd
style_header(ws2,len(hdr2))

# ===== Sheet 3: Legend =====
ws3=wb.create_sheet("Легенда")
from collections import Counter
ds=Counter(r["dataset"] for r in ents); tv=Counter((r.get("tattva") or "(писание/—)") for r in ents)
conf=Counter(r.get("confidence","") for r in ents)
def block(title,pairs,startcol=1):
    ws3.cell(row=block.row,column=startcol,value=title).font=Font(name="Arial",bold=True,size=12)
    block.row+=1
    for k,v in pairs:
        ws3.cell(row=block.row,column=startcol,value=k).font=BASE_FONT
        ws3.cell(row=block.row,column=startcol+1,value=v).font=BASE_FONT
        block.row+=1
    block.row+=1
block.row=1
ws3.cell(row=block.row,column=1,value="ИСККОН — Глоссарий имён (реестр сущностей)").font=Font(name="Arial",bold=True,size=14); block.row+=2
block("Принцип",[("Источник всего — Кришна","kṛṣṇas tu bhagavān svayam (СБ 1.3.28; Б.-с. 5.1)"),
                 ("Единый ключ","id неизменен; имена — атрибуты; меняешь в одном месте — везде")])
block("Таттва",[("vishnu-tattva","Кришна и Его экспансии/аватары"),
                ("shakti-tattva","энергии и супруги Господа"),
                ("shiva-tattva","Шива (Б.-с. 5.45)"),
                ("jiva-tattva","дживы; в т.ч. шактьявеша-аватары")])
block("Достоверность",[("verified","надёжно установлено"),("review","требует выверки")])
block("Итоги по датасетам",[(k,str(v)) for k,v in ds.items()]+[("ВСЕГО имён",str(len(ents))),("ВСЕГО связей",str(len(rels)))])
block("По таттве",[(k,str(v)) for k,v in tv.most_common()])
block("По достоверности",[(k,str(v)) for k,v in conf.most_common()])
ws3.column_dimensions["A"].width=34; ws3.column_dimensions["B"].width=60

out=os.path.join(base,"ISKCON_Glossary.xlsx")
wb.save(out)
print("entities:",len(ents),"| relations:",len(rels))
print("saved:",out)
